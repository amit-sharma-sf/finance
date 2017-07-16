from openpyxl import load_workbook
import datetime
import os.path

class Driver:
	logger = None
	file = None
	filename = None

	COLS = {'Date':1, 'Open':2, 'Close':3, 'High':4, 'Low':5}


	def __init__(self, log, filename):
		self.logger = log
		self.file = load_workbook(filename = filename)
		self.filename = filename

	def create_tabs(self, tabname):
		if(tabname in self.file.get_sheet_names()):
			return
		else:
			self.file.create_sheet(title=tabname)
			tab = self.file.get_sheet_by_name(tabname)
			tab.cell(row=1, column=self.COLS['Date']).value  = 'Date'
			tab.cell(row=1, column=self.COLS['Open']).value  = 'Open'
			tab.cell(row=1, column=self.COLS['Close']).value = 'Close'
			tab.cell(row=1, column=self.COLS['High']).value  = 'High'
			tab.cell(row=1, column=self.COLS['Low']).value   = 'Low'

	def append_data(self, instruments, tabname):
		self.insert_data(instruments, tabname, 0)

	def insert_data(self, instruments, tabname, row_count):
		if(tabname in self.file.get_sheet_names()):
			tab = self.file.get_sheet_by_name(tabname)
			self.logger.info("Found Tab: " + tabname)

			if(row_count is 0):
				row_count = 1
				for count in range(1,10000):
					if(tab.cell(row=count, column=1).value is not None):
						row_count = row_count + 1
					else:
						break
						
			self.logger.info("Found Empty Cell at index: " + row_count.__str__())
			for instr in instruments:
				tab.cell(row=row_count, column=self.COLS['Date']).value  = instr['Date']
				tab.cell(row=row_count, column=self.COLS['Open']).value  = instr['Open']
				tab.cell(row=row_count, column=self.COLS['Close']).value = instr['Close']
				tab.cell(row=row_count, column=self.COLS['High']).value  = instr['High']
				tab.cell(row=row_count, column=self.COLS['Low']).value   = instr['Low']
				row_count = row_count + 1
			
		else:
			self.logger.error("Cannot find Tab: " + tabname)

	def save(self):
		self.file.save(self.filename)
		self.logger.info("Saved File: " + self.filename)



		